# 1️⃣ WH Setup
items = ["shampoo", "soap", "brush"]
stock = [20, 30, 10]
threshold = 5
lead_time = [7, 3, 2]
max_stock = [20, 30, 10]
dispatch = [0, 0, 0]
returns = [0, 0, 0]

# 2️⃣ Record Sales
def record_sale(items, stock, threshold):
    sales = []
    cumulative = []

    for i in range(len(items)):
        sold = int(input(f"Units sold for {items[i]} (Available: {stock[i]}): "))
        while sold > stock[i]:
            print(f"⚠️ You only have {stock[i]} units of {items[i]}. Please enter a valid quantity.")
            sold = int(input(f"Re-enter units sold for {items[i]}: "))
        
        sales.append(sold)
        stock[i] -= sold
        cumulative.append(sold)

        if stock[i] <= threshold:
            print(f"⚠️ ALERT: {items[i]} has dropped to {stock[i]} units. Reorder immediately!")

        projected_weekly_sales = sold * (lead_time[i] // 1)
        projected_stock = stock[i] - projected_weekly_sales
        if projected_stock <= threshold:
            print(f"📈 Forecast: {items[i]} may drop below threshold in {lead_time[i]} days. Consider early reorder.")

    return sales, cumulative, stock

# 3️⃣ Sort by Urgency
def sort_by_stock(items, updated_stock):
    sorted_items = items[:]
    sorted_stock = updated_stock[:]

    for i in range(len(sorted_stock) - 1):
        for j in range(len(sorted_stock) - 1):
            if sorted_stock[j] > sorted_stock[j + 1]:
                sorted_stock[j], sorted_stock[j + 1] = sorted_stock[j + 1], sorted_stock[j]
                sorted_items[j], sorted_items[j + 1] = sorted_items[j + 1], sorted_items[j]

    return sorted_items, sorted_stock

# 4️⃣ Reporting
def print_report(sorted_items, sorted_stock, sales):
    print("\n📦 Restock Report")
    print("----------------------------")
    print("Item       | Stock Left")
    print("----------------------------")
    for i in range(len(sorted_items)):
        print(f"{sorted_items[i]:<10} | {sorted_stock[i]}")
    print("----------------------------")
    print(f"Total Units Sold Today: {sum(sales)}")

# 5️⃣ Excel Export
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from datetime import datetime

def export_to_excel(items, updated_stock, sales, dispatch, returns, threshold, max_stock):
    wb = Workbook()
    sheet_name = f"Report_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}"
    ws = wb.create_sheet(title=sheet_name)
    wb.remove(wb["Sheet"])  # Remove default sheet

    # Headers
    ws.append(["Item", "Stock Left", "Units Sold", "Suggested Reorder", "Dispatched", "Returned", "Reorder Flag"])

    # Bold headers
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 14

    # Data rows
    for i in range(len(items)):
        reorder_qty = max_stock[i] - updated_stock[i]
        flag = "Yes" if updated_stock[i] <= threshold else "No"
        ws.append([items[i], updated_stock[i], sales[i], reorder_qty, dispatch[i], returns[i], flag])

    # Total and timestamp
    ws.append(["", "", "", "", "", "", ""])
    ws.append(["Total", "", sum(sales), "", "", "", ""])
    ws.append(["Report Date:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

    # Highlight low stock
    low_stock_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    for row in ws.iter_rows(min_row=2, max_row=1 + len(items), min_col=2, max_col=2):
        for cell in row:
            if cell.value <= threshold:
                cell.fill = low_stock_fill

    # Save file
    filename = "smart_replenishment.xlsx"
    wb.save(filename)
    print(f"\n📁 Excel report saved as '{filename}'")

# 6️⃣ CLI Menu
def main_menu():
    print("\n📦 Smart Replenishment Menu")
    print("1. Record Sales")
    print("2. View Restock Report")
    print("3. Export to Excel")
    print("4. Exit")

    choice = input("Select an option (1–4): ")

    if choice == "1":
        return "record"
    elif choice == "2":
        return "report"
    elif choice == "3":
        return "export"
    elif choice == "4":
        print("Exiting Smart Replenishment. Goodbye!")
        exit()
    else:
        print("Invalid choice. Try again.")
        return main_menu()

# 7️⃣ Run Program
action = main_menu()

if action == "record":
    sales, cumulative, updated_stock = record_sale(items, stock, threshold)
    sorted_items, sorted_stock = sort_by_stock(items, updated_stock)
    print_report(sorted_items, sorted_stock, sales)
    export_to_excel(items, updated_stock, sales, dispatch, returns, threshold, max_stock)

elif action == "report":
    sorted_items, sorted_stock = sort_by_stock(items, stock)
    print_report(sorted_items, sorted_stock, [0]*len(items))

elif action == "export":
    export_to_excel(items, stock, [0]*len(items), dispatch, returns, threshold, max_stock)
